"""
电子发票信息提取工具
双击 EXE 或直接运行脚本，自动处理同目录下所有 PDF 并生成汇总 Excel。
"""

import sys
import re
import os
import unicodedata
from pathlib import Path
from datetime import datetime

import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ─────────────────────────── 字段提取 ───────────────────────────

def extract_page_texts(pdf_path: str) -> list:
    """按页提取 PDF 文本（一个 PDF 文件可能含多张发票，需按页分组）"""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            return [page.extract_text() or "" for page in pdf.pages]
    except Exception as e:
        print(f"  ⚠ 读取失败: {pdf_path} — {e}")
        return []


def split_invoice_texts(page_texts: list) -> list:
    """按每页的发票号码把多页 PDF 拆成独立发票的文本段。
    页上出现与前页不同的 20 位号码即视为新发票；无号码的页并入前一张（续页）。"""
    groups = []  # [发票号码 or None, [页文本...]]
    for pt in page_texts:
        m = re.search(r"发票号码[：:]\s*(\d{20})", pt)
        num = m.group(1) if m else None
        if not groups:
            groups.append([num, [pt]])
        elif num and groups[-1][0] and num != groups[-1][0]:
            groups.append([num, [pt]])
        else:
            if num and not groups[-1][0]:
                groups[-1][0] = num
            groups[-1][1].append(pt)
    return ["\n".join(g[1]) for g in groups]


def _find(pattern, text, default="", flags=0):
    m = re.search(pattern, text, flags)
    try:
        return m.group(1).strip() if m else default
    except IndexError:
        return default


def _is_num(tok: str) -> bool:
    return bool(re.fullmatch(r"-?[\d,]+(?:\.\d+)?", tok))


def _to_num(tok: str):
    try:
        return float(tok.replace(",", ""))
    except ValueError:
        return None


def _split_merged_qty_price(tok: str, amount):
    """数量与单价在 PDF 提取时可能粘连成一个数（如 280 和 87.94 变成 28087.94…），
    按「数量 × 单价 ≈ 金额」逐位试拆还原"""
    if amount is None or "." not in tok:
        return None
    int_part = tok.split(".")[0].lstrip("-")
    for i in range(1, len(int_part)):
        qty_s, price_s = tok[:i], tok[i:]
        if price_s.startswith("0") and not price_s.startswith("0."):
            continue
        qty, price = float(qty_s), float(price_s)
        if abs(round(qty * price, 2) - amount) <= 0.01:
            return qty, price
    return None


# 商品行下方的折行碎片不会包含这些内容；遇到即停止拼接
_ITEM_STOP = re.compile(r"[¥￥%]|合\s*计|价税|备\s*注|开票人|收款人|复\s*核|销\s*售|购\s*买|监\s*制|发票|开票日期|税额")


def parse_line_items(text: str) -> list:
    """解析商品明细行。行格式：*税收分类*名称 [规格] [单位] [数量 单价] 金额 税率 税额
    行尾三项结构固定，从右往左解析；数量/单价/单位按剩余 token 尽力还原。
    名称/规格过长时 PDF 会折行，紧随其后的碎片行按 全中文→名称、含字母数字→规格 拼回。"""
    lines = text.split("\n")
    items = []
    i = 0
    while i < len(lines):
        m = re.match(r"^\s*\*([^*\n]+)\*\s*(.+)$", lines[i])
        i += 1
        if not m:
            continue
        category, rest = m.group(1).strip(), m.group(2).strip()
        tokens = rest.split()
        if len(tokens) < 3:
            continue
        rate = tokens[-2]
        if not (rate.endswith("%") or rate in ("免税", "不征税", "***")):
            continue
        if not (_is_num(tokens[-1]) and _is_num(tokens[-3])):
            continue
        tax_amount = _to_num(tokens[-1])
        amount = _to_num(tokens[-3])

        mid = tokens[:-3]
        name = mid.pop(0) if mid else ""
        qty = price = None
        unit = ""
        if mid and _is_num(mid[-1]):
            price_tok = mid.pop()
            if mid and _is_num(mid[-1]):
                qty = _to_num(mid.pop())
                price = _to_num(price_tok)
            else:
                merged = _split_merged_qty_price(price_tok.replace(",", ""), amount)
                if merged:
                    qty, price = merged
                else:
                    price = _to_num(price_tok)
        if mid and not _is_num(mid[-1]) and re.fullmatch(r"[一-鿿]{1,3}", mid[-1]):
            unit = mid.pop()
        spec = " ".join(mid)

        # 折行恢复：最多向下拼 3 行碎片。
        # 名称列窄，折行出的名称碎片必然短，只认第一个 ≤6 字的纯中文碎片；其余归规格
        taken = 0
        name_extended = False
        while i < len(lines) and taken < 3:
            cont = lines[i].strip()
            if not cont or cont.startswith("*") or _ITEM_STOP.search(cont):
                break
            ctoks = cont.split()
            if len(ctoks) > 3 or any(_is_num(t) for t in ctoks):
                break
            for t in ctoks:
                if not name_extended and re.fullmatch(r"[一-鿿]{1,6}", t):
                    name += t
                    name_extended = True
                else:
                    spec = f"{spec} {t}".strip()
            i += 1
            taken += 1

        items.append({
            "税收分类": category,
            "商品名称": name,
            "规格型号": spec,
            "单位": unit,
            "数量": qty,
            "单价": price,
            "金额": amount,
            "税率": rate,
            "税额": tax_amount,
        })
    return items


def parse_invoice(text: str, filename: str) -> dict:
    """从发票文本中解析关键字段（兼容增值税电子发票 & 全电发票）"""
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    # 部分 PDF 内嵌字体将"子"等字映射到康熙部首区（如"电⼦发票"），归一化为常规汉字
    text = "".join(
        unicodedata.normalize("NFKC", ch) if "⺀" <= ch <= "⿟" else ch
        for ch in text
    )

    # ── 发票类型 ──────────────────────────────────────────────────
    type_map = {
        "全电发票": "全电发票",
        "增值税电子专用发票": "增值税电子专用发票",
        "增值税专用发票": "增值税专用发票",
        "增值税电子普通发票": "增值税电子普通发票",
        "增值税普通发票": "增值税普通发票",
        "电子发票": "电子发票（普通发票）",
    }
    invoice_type = "未知"
    for kw, label in type_map.items():
        if kw in text:
            invoice_type = label
            break

    # ── 发票代码 / 号码 ───────────────────────────────────────────
    # 旧版：发票代码（10/12位）+ 发票号码（8位）；数电票：仅20位发票号码、无发票代码
    code = _find(r"发票代码[：:]\s*(\d{10,12})", text)
    # 上限放到20位贪婪匹配，避免20位号码被旧版8位规则截断
    number = _find(r"发票号码[：:]\s*(\d{6,20})", text)
    if not code and not number:
        # 尝试无标签的 20 位号码
        m = re.search(r"\b(\d{20})\b", text)
        if m:
            number = m.group(1)

    # ── 开票日期 ──────────────────────────────────────────────────
    date = _find(r"开票日期[：:]\s*(\d{4}年\s*\d{1,2}月\s*\d{1,2}日)", text)
    if not date:
        date = _find(r"开票日期[：:]\s*(\d{4}-\d{2}-\d{2})", text)
    if not date:
        date = _find(r"(\d{4}年\s*\d{1,2}月\s*\d{1,2}日)", text)
    if date:
        date = re.sub(r'\s+', '', date)  # 统一去除年月日间的空格

    # ── 购买方 ────────────────────────────────────────────────────
    buyer = _find(r"购\s*买\s*方\s*名\s*称[：:]\s*(.+)", text)
    if not buyer:
        # 双栏布局：同行 "购 名称：xxx 销 名称：yyy"，止于"销 名称："
        buyer = _find(r"购.{0,10}名\s*称[：:]\s*(.+?)(?=\s+销\s+名\s*称|\s+纳税人|\s*\n|$)", text)
    buyer_tax_id = _find(r"购买方.{0,2}纳税人识别号[：:]\s*([A-Z0-9]{15,20})", text)
    if not buyer_tax_id:
        buyer_tax_id = _find(r"统一社会信用代码/纳税人识别号[：:]\s*([A-Z0-9]{15,20})", text)

    # ── 销售方 ────────────────────────────────────────────────────
    seller = _find(r"销\s*售\s*方\s*名\s*称[：:]\s*(.+)", text)
    if not seller:
        seller = _find(r"销售方\s*名称[：:]\s*(.+)", text)
    if not seller:
        # 双栏布局：同行末尾 "销 名称：yyy"
        seller = _find(r"销\s+名\s*称[：:]\s*(.+?)(?:\s*\n|$)", text)
    seller_tax_id = _find(r"销.{0,4}纳税人识别号[：:]\s*([A-Z0-9]{15,20})", text)
    if not seller_tax_id:
        # 双栏布局：统一社会信用代码出现两次，第二个是销售方
        matches = re.findall(r"统一社会信用代码/纳税人识别号[：:]\s*([A-Z0-9]{15,20})", text)
        if len(matches) >= 2:
            seller_tax_id = matches[1]

    if not buyer or not seller:
        # 无标签双栏布局：公司名出现在正文区，两个名称同行
        _co = r'[一-鿿A-Za-z（）()·]{4,}?(?:有限公司|有限责任公司|个体工商户|经营部(?:（个体工商户）)?|集团|有限合伙企业)'
        m = re.search(fr'^({_co})\s+({_co})', text, re.MULTILINE)
        if m:
            if not buyer:
                buyer = m.group(1).strip()
            if not seller:
                seller = m.group(2).strip()
        # 对应格式的税号：两个税号同行
        if not buyer_tax_id or not seller_tax_id:
            id_pairs = re.findall(r'^([A-Z0-9]{15,20})\s+([A-Z0-9]{15,20})\s*$', text, re.MULTILINE)
            if id_pairs:
                if not buyer_tax_id:
                    buyer_tax_id = id_pairs[0][0]
                if not seller_tax_id:
                    seller_tax_id = id_pairs[0][1]

    # PDF 字符间距会让提取的公司名混入空格（如"广 东 能 环…"），统一去除
    if buyer:
        buyer = re.sub(r"\s+", "", buyer)
    if seller:
        seller = re.sub(r"\s+", "", seller)

    # ── 金额 ──────────────────────────────────────────────────────
    # 合计金额（不含税）
    subtotal = _find(r"合\s*计\s*[¥￥]?\s*([\d,]+\.?\d*)\s*[¥￥]?\s*[\d,]+\.?\d*", text)
    if not subtotal:
        subtotal = _find(r"[¥￥]\s*([\d,]+\.\d{2})\s+[¥￥]", text)

    # 合计税额（用 \s* 代替 .*? 以支持合计与金额跨行的格式）
    tax = _find(r"合\s*计\s*[¥￥]?\s*[\d,]+\.?\d*\s*[¥￥]\s*([\d,]+\.?\d*)", text)

    # 价税合计（小写）；兼容"价税合计（大写）xxx（小写）¥yyy"格式
    total = _find(r"(?:价税合计|合计金额).*?[（(]小写[）)]\s*[¥￥]\s*([\d,]+\.?\d*)", text)
    if not total:
        m = re.search(r"[¥￥]\s*([\d,]+\.\d{2})\s*$", text, re.MULTILINE)
        total = m.group(1).strip() if m else ""
    if not total:
        # 全电发票
        total = _find(r"合\s*计\s*([\d,]+\.\d{2})", text)

    # ── 备注 ──────────────────────────────────────────────────────
    remark = _find(r"备\s*注[：:]\s*(.+?)(?:\n|$)", text)

    # ── 商品/服务明细 ─────────────────────────────────────────────
    line_items = parse_line_items(text)
    if line_items:
        item = line_items[0]["商品名称"]
        if len(line_items) > 1:
            item = f"{item}（等{len(line_items)}项）"
    else:
        item = _find(r"\*[^*\n]+\*([^\n]+)", text)  # 常见格式：*类别*商品名
        if not item:
            item = _find(r"货物或应税劳务.*?\n(.+?)(?:\s+\d|\s+\*)", text)

    return {
        "_明细":          line_items,
        "文件名":        filename,
        "发票类型":       invoice_type,
        "发票代码":       code,
        "发票号码":       number,
        "开票日期":       date,
        "购买方名称":     buyer,
        "购买方税号":     buyer_tax_id,
        "销售方名称":     seller,
        "销售方税号":     seller_tax_id,
        "合计金额(不含税)": subtotal,
        "合计税额":       tax,
        "价税合计":       total,
        "商品/服务":      item,
        "备注":           remark,
    }


# ─────────────────────────── Excel 输出 ───────────────────────────

HEADERS = [
    "文件名", "发票类型", "发票代码", "发票号码", "开票日期",
    "购买方名称", "购买方税号", "销售方名称", "销售方税号",
    "合计金额(不含税)", "合计税额", "价税合计", "商品/服务", "备注",
]

NUMERIC_COLS = {"合计金额(不含税)", "合计税额", "价税合计"}

COL_WIDTHS = {
    "文件名": 30, "发票类型": 18, "发票代码": 16, "发票号码": 14,
    "开票日期": 16, "购买方名称": 28, "购买方税号": 22,
    "销售方名称": 28, "销售方税号": 22,
    "合计金额(不含税)": 16, "合计税额": 14, "价税合计": 14,
    "商品/服务": 28, "备注": 30,
}

HEADER_BG = "1F4E79"   # 深蓝
HEADER_FG = "FFFFFF"
ALT_ROW_BG = "DCE6F1"  # 浅蓝


def write_excel(records: list, output_path: str, pdf_folder: Path = None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "发票汇总"

    # ── 标题行 ────────────────────────────────────────────────────
    header_font   = Font(name="微软雅黑", bold=True, color=HEADER_FG, size=11)
    header_fill   = PatternFill("solid", fgColor=HEADER_BG)
    center_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.row_dimensions[1].height = 28
    for col_idx, col_name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    # ── 数据行 ────────────────────────────────────────────────────
    data_font      = Font(name="微软雅黑", size=10)
    link_font      = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
    for row_idx, record in enumerate(records, start=2):
        is_alt = row_idx % 2 == 0
        row_fill = PatternFill("solid", fgColor=ALT_ROW_BG) if is_alt else None
        ws.row_dimensions[row_idx].height = 20

        for col_idx, col_name in enumerate(HEADERS, start=1):
            value = record.get(col_name, "")
            if col_name in NUMERIC_COLS and value:
                try:
                    value = float(str(value).replace(",", ""))
                except ValueError:
                    pass
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if cell.data_type == "f":
                # PDF 提取的文本若以 = 开头，openpyxl 会当作公式写入，强制按文本处理防止公式注入
                cell.data_type = "s"
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if col_name in ("发票代码","发票号码","开票日期","合计金额(不含税)","合计税额","价税合计") else "left",
                vertical="center",
                wrap_text=True,
            )
            if col_name == "文件名" and pdf_folder and value:
                cell.hyperlink = str(value)  # 相对路径，仅文件名
                cell.font = link_font
            else:
                cell.font = data_font
            if row_fill:
                cell.fill = row_fill

    # ── 列宽 ──────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(HEADERS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = COL_WIDTHS.get(col_name, 18)

    # ── 冻结首行 ──────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── 商品明细 Sheet（每条商品一行）──────────────────────────────
    DETAIL_HEADERS = ["文件名", "发票号码", "税收分类", "商品名称", "规格型号",
                      "单位", "数量", "单价", "金额", "税率", "税额"]
    DETAIL_WIDTHS = {"文件名": 30, "发票号码": 22, "税收分类": 16, "商品名称": 20,
                     "规格型号": 26, "单位": 8, "数量": 10, "单价": 12,
                     "金额": 12, "税率": 8, "税额": 12}
    DETAIL_NUM_FMT = {"数量": "General", "单价": "#,##0.00##", "金额": "#,##0.00", "税额": "#,##0.00"}

    wsd = wb.create_sheet("商品明细")
    wsd.row_dimensions[1].height = 28
    for col_idx, col_name in enumerate(DETAIL_HEADERS, start=1):
        cell = wsd.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
        wsd.column_dimensions[cell.column_letter].width = DETAIL_WIDTHS.get(col_name, 14)

    detail_row = 2
    for record in records:
        for li in record.get("_明细", []):
            is_alt = detail_row % 2 == 0
            row_fill = PatternFill("solid", fgColor=ALT_ROW_BG) if is_alt else None
            values = {"文件名": record.get("文件名", ""), "发票号码": record.get("发票号码", ""), **li}
            for col_idx, col_name in enumerate(DETAIL_HEADERS, start=1):
                value = values.get(col_name)
                cell = wsd.cell(row=detail_row, column=col_idx, value=value)
                if cell.data_type == "f":
                    cell.data_type = "s"  # 防公式注入，同汇总表
                cell.font = data_font
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="left" if col_name in ("文件名", "商品名称", "规格型号") else "center",
                    vertical="center", wrap_text=True,
                )
                if col_name in DETAIL_NUM_FMT and isinstance(value, float):
                    cell.number_format = DETAIL_NUM_FMT[col_name]
                if row_fill:
                    cell.fill = row_fill
            detail_row += 1
    wsd.freeze_panes = "A2"

    # ── 汇总 Sheet ────────────────────────────────────────────────
    ws2 = wb.create_sheet("统计")
    ws2["A1"] = "统计项"
    ws2["B1"] = "数值"
    for cell in [ws2["A1"], ws2["B1"]]:
        cell.font = Font(name="微软雅黑", bold=True, color=HEADER_FG, size=11)
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    stats = [
        ("发票总张数", "=COUNTA('发票汇总'!A2:A10000)"),
        ("价税合计总金额", "=SUMPRODUCT(IFERROR(VALUE(SUBSTITUTE('发票汇总'!L2:L10000,\",\",\"\")),0))"),
    ]
    for r, (label, formula) in enumerate(stats, start=2):
        ws2.cell(row=r, column=1, value=label).font = Font(name="微软雅黑", size=10)
        ws2.cell(row=r, column=2, value=formula).font = Font(name="微软雅黑", size=10)
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 22

    wb.save(output_path)


# ─────────────────────────── 主流程 ───────────────────────────────

def pause_if_frozen():
    """EXE 双击运行时控制台会立即关闭，结束前暂停以便用户查看结果"""
    if getattr(sys, "frozen", False):
        input("\n按回车键退出...")


def get_base_dir() -> Path:
    """返回脚本/EXE 所在目录，兼容 PyInstaller 打包和直接运行两种模式。"""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent  # PyInstaller EXE 模式
    return Path(__file__).resolve().parent  # 直接运行脚本模式


def main():
    folder = Path(sys.argv[1]) if len(sys.argv) > 1 else get_base_dir()

    if not folder.exists():
        print(f"❌ 文件夹不存在: {folder.resolve()}")
        pause_if_frozen()
        sys.exit(1)

    # 只扫当前层，不递归子文件夹；按后缀小写匹配，兼容 .PDF 等大写扩展名
    pdf_files = sorted(p for p in folder.iterdir() if p.is_file() and p.suffix.lower() == ".pdf")
    if not pdf_files:
        print(f"⚠ 在 {folder.resolve()} 中未找到 PDF 文件")
        pause_if_frozen()
        sys.exit(1)

    print(f"📂 扫描文件夹: {folder.resolve()}")
    print(f"🔍 找到 {len(pdf_files)} 个 PDF 文件\n")

    records = []
    failed = []

    for pdf_path in pdf_files:
        print(f"  → 处理: {pdf_path.name}")
        page_texts = extract_page_texts(str(pdf_path))
        if not "".join(page_texts).strip():
            print(f"    ⚠ 文本为空，可能是扫描件或加密 PDF")
            failed.append(pdf_path.name)
            continue
        # 一个 PDF 文件可能含多张发票（按页上的发票号码拆分），每张一条记录
        segments = split_invoice_texts(page_texts)
        if len(segments) > 1:
            print(f"    ℹ 该文件含 {len(segments)} 张发票")
        for seg in segments:
            records.append(parse_invoice(seg, pdf_path.name))

    if not records:
        print("\n❌ 没有成功解析任何发票")
        pause_if_frozen()
        sys.exit(1)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = folder / f"发票汇总_{timestamp}.xlsx"
    write_excel(records, str(output_path), pdf_folder=folder)

    print(f"\n✅ 完成！共处理 {len(records)} 张发票")
    if failed:
        print(f"⚠ 失败 {len(failed)} 个: {', '.join(failed)}")
    print(f"📊 输出文件: {output_path.resolve()}")
    pause_if_frozen()


if __name__ == "__main__":
    # 控制台编码不支持 emoji 时（如重定向到 GBK 文件）降级为替换字符，避免直接崩溃
    for stream in (sys.stdout, sys.stderr):
        if stream and hasattr(stream, "reconfigure"):
            stream.reconfigure(errors="replace")
    try:
        main()
    except SystemExit:
        raise
    except Exception:
        import traceback
        traceback.print_exc()
        pause_if_frozen()
        sys.exit(1)
